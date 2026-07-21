from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path


class ResourceLimitExceeded(RuntimeError):
    pass


class ResourceInspectionFailed(RuntimeError):
    pass


@dataclass(frozen=True)
class DocumentLimits:
    max_bytes: int
    max_pdf_pages: int
    max_image_width: int
    max_image_height: int
    max_sheet_rows: int
    max_sheet_columns: int
    max_nonempty_cells: int
    max_archive_uncompressed_bytes: int = 500 * 1024 * 1024
    max_archive_ratio: int = 100


def inspect_resources(path: Path, limits: DocumentLimits) -> None:
    try:
        if path.stat().st_size > limits.max_bytes:
            raise ResourceLimitExceeded("文件大小超过限制")
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            _inspect_pdf(path, limits)
        elif suffix in {".png", ".jpg", ".jpeg", ".tiff", ".bmp"}:
            _inspect_image(path, limits)
        elif suffix == ".xlsx":
            _inspect_spreadsheet(path, limits)
        if suffix in {".docx", ".xlsx", ".pptx"}:
            _inspect_archive(path, limits)
    except ResourceLimitExceeded:
        raise
    except Exception as exc:
        raise ResourceInspectionFailed("文档资源检查失败") from exc


def _inspect_pdf(path: Path, limits: DocumentLimits) -> None:
    from pypdf import PdfReader

    reader = PdfReader(path, strict=True)
    if len(reader.pages) > limits.max_pdf_pages:
        raise ResourceLimitExceeded("PDF 页数超过限制")


def _inspect_image(path: Path, limits: DocumentLimits) -> None:
    from PIL import Image

    with Image.open(path) as image:
        width, height = image.size
        if width > limits.max_image_width or height > limits.max_image_height:
            raise ResourceLimitExceeded("图像像素尺寸超过限制")


def _inspect_spreadsheet(path: Path, limits: DocumentLimits) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        nonempty = 0
        for sheet in workbook.worksheets:
            if sheet.max_row > limits.max_sheet_rows:
                raise ResourceLimitExceeded("工作表行数超过限制")
            if sheet.max_column > limits.max_sheet_columns:
                raise ResourceLimitExceeded("工作表列数超过限制")
            for row in sheet.iter_rows():
                nonempty += sum(cell.value is not None for cell in row)
                if nonempty > limits.max_nonempty_cells:
                    raise ResourceLimitExceeded("工作表非空单元格超过限制")
    finally:
        workbook.close()


def _inspect_archive(path: Path, limits: DocumentLimits) -> None:
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        uncompressed = sum(item.file_size for item in infos)
        compressed = max(1, sum(item.compress_size for item in infos))
        if uncompressed > limits.max_archive_uncompressed_bytes:
            raise ResourceLimitExceeded("解压后文件大小超过限制")
        if uncompressed / compressed > limits.max_archive_ratio:
            raise ResourceLimitExceeded("压缩比超过限制")
