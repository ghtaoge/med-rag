"""Excel 加载器。openpyxl 逐行提取表格数据。"""

from __future__ import annotations

from pathlib import Path

from app.documents.loader.base import DocumentLoader


class ExcelLoader(DocumentLoader):
    """Excel(.xlsx) 加载器。"""

    def load(self, file_path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Excel 加载需要 openpyxl，请执行: pip install openpyxl")

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("| " + " | ".join(cells) + " |")
            if rows:
                parts.append(f"# 工作表: {sheet.title}\n\n" + "\n".join(rows))

        wb.close()
        return "\n\n".join(parts)
